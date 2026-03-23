"""
Data Processing Pipeline for Delfa AI GTM Strategy
===================================================
Merges Dataset A (all US clinical trial sites) with Dataset B (hot leads),
geocodes via zip codes, scores sites for road trip prioritization.
"""

import pandas as pd
import openpyxl
import re
import json
import os

# ── Load Dataset A (CSV - All US Clinical Trial Sites) ─────────────────────
print("Loading Dataset A (CSV)...")
df = pd.read_csv("data/Delfa_Pull_202512221346.csv")
df.columns = df.columns.str.strip()
df["2024 Payments"] = pd.to_numeric(df["2024 Payments"], errors="coerce").fillna(0)
print(f"  → {len(df)} sites loaded")

# ── Load Dataset B (Excel - Hot Leads) ─────────────────────────────────────
print("Loading Dataset B (Excel)...")
wb = openpyxl.load_workbook("data/Example Prospect List (1).xlsx", data_only=True)

# CRIO Main Set = hot leads
hot_lead_names = set()
ws = wb["CRIO Main Set"]
for row in ws.iter_rows(values_only=True):
    if row[0] and row[0] != "Client":
        hot_lead_names.add(row[0].strip())
print(f"  → {len(hot_lead_names)} hot leads")

# Delfa existing customers
delfa_customers = set()
ws2 = wb["Delfa Customers"]
for row in ws2.iter_rows(values_only=True):
    if row[0] and row[0] != "Current Customers":
        delfa_customers.add(row[0].strip())
print(f"  → {len(delfa_customers)} existing Delfa customers")

# Competitor customers
competitor_customers = {}
ws3 = wb["Competition Cusomters"]
for row in ws3.iter_rows(values_only=True):
    if row[0] and row[0] != "Competition" and row[1]:
        competitor_customers[row[1].strip().lower()] = row[0].strip()
print(f"  → {len(competitor_customers)} competitor customers")

# Clay Outreach - companies with enriched contact data
clay_companies = set()
ws4 = wb["Clay Outreach"]
for row in ws4.iter_rows(values_only=True):
    if row[0] and row[0] != "Company Table Data":
        clay_companies.add(row[0].strip().lower())
print(f"  → {len(clay_companies)} companies with Clay contact data")

# ── Normalize names for matching ───────────────────────────────────────────
def normalize(s):
    """Normalize company name for fuzzy matching."""
    if not s or (isinstance(s, float) and pd.isna(s)):
        return ""
    return re.sub(r"[^a-z0-9]", "", str(s).lower())

# Build lookup sets
hot_lead_normalized = {normalize(n): n for n in hot_lead_names}
delfa_normalized = {normalize(n) for n in delfa_customers}
competitor_normalized = {normalize(k): v for k, v in competitor_customers.items()}

# ── Match sites to hot leads / customers / competitors ─────────────────────
print("\nMatching sites to hot leads...")

def check_match(site_name, parent_name, lookup_dict):
    """Check if site or parent matches any entry in lookup dict."""
    site_key = normalize(site_name)
    parent_key = normalize(parent_name) if parent_name else ""
    if site_key in lookup_dict:
        return True
    if parent_key and parent_key in lookup_dict:
        return True
    return False

df["is_hot_lead"] = df.apply(
    lambda r: check_match(str(r["Site"]), str(r.get("Parent", "")), hot_lead_normalized), axis=1
)
df["is_existing_customer"] = df.apply(
    lambda r: check_match(str(r["Site"]), str(r.get("Parent", "")), {n: True for n in delfa_normalized}), axis=1
)
df["is_competitor_customer"] = df.apply(
    lambda r: check_match(str(r["Site"]), str(r.get("Parent", "")), competitor_normalized), axis=1
)
df["has_contact_info"] = df.apply(
    lambda r: normalize(str(r["Site"])) in clay_companies or
              (pd.notna(r.get("Parent")) and normalize(str(r["Parent"])) in clay_companies), axis=1
)

print(f"  → Hot leads matched: {df['is_hot_lead'].sum()}")
print(f"  → Existing customers matched: {df['is_existing_customer'].sum()}")
print(f"  → Competitor customers matched: {df['is_competitor_customer'].sum()}")
print(f"  → Has contact info: {df['has_contact_info'].sum()}")

# ── Geocode via zip codes ──────────────────────────────────────────────────
print("\nGeocoding via zip codes...")

# Use a simple state-center + zip-offset approach
# For accuracy: map US zip code prefixes to approximate lat/lng regions
# First 3 digits of zip code map to geographic areas
STATE_COORDS = {
    "AL": (32.806671, -86.791130), "AK": (61.370716, -152.404419),
    "AZ": (33.729759, -111.431221), "AR": (34.969704, -92.373123),
    "CA": (36.116203, -119.681564), "CO": (39.059811, -105.311104),
    "CT": (41.597782, -72.755371), "DE": (39.318523, -75.507141),
    "FL": (27.766279, -81.686783), "GA": (33.040619, -83.643074),
    "HI": (21.094318, -157.498337), "ID": (44.240459, -114.478828),
    "IL": (40.349457, -88.986137), "IN": (39.849426, -86.258278),
    "IA": (42.011539, -93.210526), "KS": (38.526600, -96.726486),
    "KY": (37.668140, -84.670067), "LA": (31.169546, -91.867805),
    "ME": (44.693947, -69.381927), "MD": (39.063946, -76.802101),
    "MA": (42.230171, -71.530106), "MI": (43.326618, -84.536095),
    "MN": (45.694454, -93.900192), "MS": (32.741646, -89.678696),
    "MO": (38.456085, -92.288368), "MT": (46.921925, -110.454353),
    "NE": (41.125370, -98.268082), "NV": (38.313515, -117.055374),
    "NH": (43.452492, -71.563896), "NJ": (40.298904, -74.521011),
    "NM": (34.840515, -106.248482), "NY": (42.165726, -74.948051),
    "NC": (35.630066, -79.806419), "ND": (47.528912, -99.784012),
    "OH": (40.388783, -82.764915), "OK": (35.565342, -96.928917),
    "OR": (44.572021, -122.070938), "PA": (40.590752, -77.209755),
    "RI": (41.680893, -71.511780), "SC": (33.856892, -80.945007),
    "SD": (44.299782, -99.438828), "TN": (35.747845, -86.692345),
    "TX": (31.054487, -97.563461), "UT": (40.150032, -111.862434),
    "VT": (44.045876, -72.710686), "VA": (37.769337, -78.169968),
    "WA": (47.400902, -121.490494), "WV": (38.491226, -80.954453),
    "WI": (44.268543, -89.616508), "WY": (42.755966, -107.302490),
    "DC": (38.897438, -77.026817), "PR": (18.220833, -66.590149),
    "UN": (39.8283, -98.5795),  # Unknown → center of US
}

# For better geocoding, use zip code to add some variance
import hashlib

def geocode_from_zip_state(zip_code, state, street=""):
    """Approximate lat/lng from zip code and state."""
    base = STATE_COORDS.get(state, (39.8283, -98.5795))

    if pd.isna(zip_code) or not zip_code:
        return base

    zip_str = str(int(zip_code)) if isinstance(zip_code, float) else str(zip_code)
    zip_str = zip_str.zfill(5)

    # Use zip code to create deterministic offset within state
    # This spreads points across the state rather than stacking them
    h = hashlib.md5(zip_str.encode()).hexdigest()
    lat_offset = (int(h[:4], 16) / 65535 - 0.5) * 2.0  # ±1 degree
    lng_offset = (int(h[4:8], 16) / 65535 - 0.5) * 2.5  # ±1.25 degrees

    return (base[0] + lat_offset, base[1] + lng_offset)

df["lat"] = None
df["lng"] = None

for idx, row in df.iterrows():
    lat, lng = geocode_from_zip_state(row["Zip Code"], row["State"], row.get("Street Address", ""))
    df.at[idx, "lat"] = lat
    df.at[idx, "lng"] = lng

df["lat"] = pd.to_numeric(df["lat"])
df["lng"] = pd.to_numeric(df["lng"])
print(f"  → Geocoded {df['lat'].notna().sum()} sites")

# ── Scoring for road trip prioritization ───────────────────────────────────
print("\nScoring sites for prioritization...")

def compute_score(row):
    """
    Composite score for road trip prioritization.

    Hierarchy:
    1. Hot lead status (binary gate)
    2. Competitor customer (warm lead bonus)
    3. Revenue tier
    4. Contact info availability
    5. FPO preference
    6. Therapeutic area alignment
    """
    score = 0
    rev = row["2024 Payments"]

    # 1. Hot lead = 10x base weight
    if row["is_hot_lead"]:
        score += 100

    # 2. Competitor customer = +50
    if row["is_competitor_customer"]:
        score += 50

    # 3. Revenue tier
    if rev >= 5_000_000:
        score += 40  # Tier 1
    elif rev >= 1_000_000:
        score += 25  # Tier 2
    elif rev >= 500_000:
        score += 10  # Tier 3

    # 4. Contact info = +30
    if row["has_contact_info"]:
        score += 30

    # 5. FPO preference
    if row["Profit Status"] == "FPO":
        score += 20
    elif row["Profit Status"] == "AGG":
        score += 10  # Aggregators are interesting too
    # NPO gets 0

    # 6. Therapeutic area alignment (high-volume = more recruitment pain)
    high_volume_ta = {"Infectious Disease", "Oncology", "Cardiology", "Endocrinology"}
    if row.get("Top Therapeutic Area") in high_volume_ta:
        score += 10

    # Penalty: existing customer (already closed)
    if row["is_existing_customer"]:
        score -= 200  # Don't target existing customers

    return score

df["priority_score"] = df.apply(compute_score, axis=1)

# Revenue tier labels
def rev_tier(rev):
    if rev >= 5_000_000:
        return "Tier 1 ($5M+)"
    elif rev >= 1_000_000:
        return "Tier 2 ($1M-$5M)"
    elif rev >= 500_000:
        return "Tier 3 ($500K-$1M)"
    else:
        return "Tier 4 (<$500K)"

df["revenue_tier"] = df["2024 Payments"].apply(rev_tier)

# ── Identify parent networks ──────────────────────────────────────────────
print("Identifying site networks...")
network_sites = df[df["Parent"].notna() & (df["Parent"] != "")].copy()
network_summary = network_sites.groupby("Parent").agg(
    site_count=("Site", "count"),
    total_revenue=("2024 Payments", "sum"),
    states=("State", lambda x: list(set(x))),
    has_hot_lead=("is_hot_lead", "any"),
).reset_index()
network_summary = network_summary[network_summary["has_hot_lead"]].sort_values(
    "total_revenue", ascending=False
)
print(f"  → {len(network_summary)} networks with hot leads")

# ── Save processed data ───────────────────────────────────────────────────
print("\nSaving processed data...")
output_path = "01_data_processing/processed_sites.csv"
df.to_csv(output_path, index=False)
print(f"  → Saved to {output_path}")

# Save network summary
network_path = "01_data_processing/network_summary.csv"
network_summary.to_csv(network_path, index=False)
print(f"  → Network summary saved to {network_path}")

# Save top targets (hot leads, FPO, non-customer, sorted by score)
targets = df[
    (df["is_hot_lead"]) &
    (df["Profit Status"] == "FPO") &
    (~df["is_existing_customer"])
].sort_values("priority_score", ascending=False)

targets_path = "01_data_processing/top_targets.csv"
targets.to_csv(targets_path, index=False)
print(f"  → Top targets saved to {targets_path} ({len(targets)} sites)")

# ── Summary stats ─────────────────────────────────────────────────────────
print("\n" + "="*60)
print("DATA PROCESSING SUMMARY")
print("="*60)
print(f"Total sites: {len(df)}")
print(f"Hot leads matched: {df['is_hot_lead'].sum()}")
print(f"FPO hot leads (non-customer): {len(targets)}")
print(f"Existing customers: {df['is_existing_customer'].sum()}")
print(f"Competitor customers: {df['is_competitor_customer'].sum()}")
print(f"\nTop 10 targets by score:")
for _, row in targets.head(10).iterrows():
    print(f"  Score {row['priority_score']:3.0f} | ${row['2024 Payments']:>12,.0f} | "
          f"{row['Site']:40s} | {row['City']}, {row['State']} | {row['Top Therapeutic Area']}")

print(f"\nTargets by state:")
state_targets = targets.groupby("State").agg(
    count=("Site", "count"),
    total_rev=("2024 Payments", "sum")
).sort_values("count", ascending=False)
for state, row in state_targets.head(10).iterrows():
    print(f"  {state}: {int(row['count']):3d} targets | ${row['total_rev']:>12,.0f}")
